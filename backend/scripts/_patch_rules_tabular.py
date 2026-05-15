"""
Rebuild recommendation_rules_simple sheet with ALL 30 rules from recommendation_rules.json.
sensor_rules(21) + maintenance_rules(7) + fallback_rules(2) = 30 total.
Engine reads this sheet — all rules must be here.
"""
import json as _json
import sys
from pathlib import Path
from openpyxl import load_workbook

DATA_DIR = Path(__file__).resolve().parent.parent / "data"

XLSX = max(
    (p for p in DATA_DIR.glob("sample_dataset*.xlsx")
     if "tabular" not in p.name and "with_rules" not in p.name),
    key=lambda p: p.stat().st_mtime,
)


def cj(obj):
    return _json.dumps(obj, separators=(",", ":"))


# ── 14-column schema read by recommendation_engine._load_rules_from_dataset ────
COLS = [
    "rule_id", "variable", "seuil_min", "seuil_max", "mode",
    "type_risque", "niveau", "message", "component",
    "suggestion_priority", "suggestion_title", "suggestion_message",
    "ruleset", "condition_json",
]

ENGINE_CONFIG_COLS = ["key", "value_json"]
ENGINE_CONFIG_ROWS = [
    ("score_floor.critical_count_ge_2", "85"),
    ("score_floor.critical_any", "70"),
    ("score_floor.warning_any", "40"),
    ("severity_promotion.critical_min_score", "70"),
    ("severity_promotion.warning_min_score", "40"),
    ("defaults.normal_suggestion.priority", cj("low")),
    ("defaults.normal_suggestion.title", cj("Etat normal")),
    ("defaults.normal_suggestion.message", cj("Aucune anomalie detectee actuellement.")),
]

# (rule_id, variable, seuil_min, seuil_max, mode,
#  type_risque, niveau, message, component,
#  sug_priority, sug_title, sug_message, ruleset, condition_json)
ROWS = [
    # ── sensor_rules (21) ─────────────────────────────────────────────────────
    ("battery_critical_low",        "battery_voltage", 11.5, None, "lt_min",
     "battery","critical","Batterie critique detectee","battery_system",
     "high","Batterie / alternateur",
     "Controler immediatement la batterie et le circuit de charge.",
     "sensor_rules", None),

    ("battery_warning_low",         "battery_voltage", 12.0, None, "lt_min",
     "battery","warning","Batterie en baisse detectee","battery_system",
     "medium","Controle batterie",
     "Planifier une verification batterie dans les prochaines 48h.",
     "sensor_rules", None),

    ("battery_warning_alt_suspect", "battery_voltage", None, None, None,
     "battery","warning","Tension basse moteur allume - alternateur suspect","",
     "medium","Test alternateur",
     "Faire diagnostiquer l'alternateur ; la batterie ne se recharge pas correctement.",
     "sensor_rules",
     cj({"all":[{"field":"battery_voltage","op":"gte","value":12.0},
                {"field":"battery_voltage","op":"lt","value":12.5},
                {"field":"rpm","op":"gt","value":500}]})),

    ("battery_critical_overvoltage","battery_voltage", None, 15.8, "gt_max",
     "battery","critical","Surtension batterie critique detectee","",
     "high","Systeme electrique",
     "Verifier l'alternateur et le systeme de charge. Risque de dommage electrique.",
     "sensor_rules", None),

    ("battery_warning_overvoltage", "battery_voltage", None, 15.5, "gt_max",
     "battery","warning","Surtension batterie detectee","",
     "medium","Test alternateur",
     "Faire verifier le systeme de charge et l'alternateur rapidement.",
     "sensor_rules", None),

    ("temp_critical",  "engine_temp", None, 106, "gt_max",
     "cooling","critical","Surchauffe moteur detectee","cooling_system",
     "high","Systeme de refroidissement",
     "Verifier liquide de refroidissement, ventilateur et radiateur.",
     "sensor_rules", None),

    ("temp_warning",   "engine_temp", None, None, None,
     "cooling","warning","Temperature moteur elevee","cooling_system",
     "medium","Inspection moteur",
     "Prevoir une inspection du systeme de refroidissement.",
     "sensor_rules",
     cj({"field":"engine_temp","op":"gte","value":99})),

    ("fuel_critical",  "fuel_level", 5, None, "lt_min",
     "fuel","critical","Niveau de carburant critique","fuel_system",
     "high","Ravitaillement urgent",
     "Ravitailler immediatement - risque d'arret moteur.",
     "sensor_rules", None),

    ("fuel_warning",   "fuel_level", 12, None, "lt_min",
     "fuel","warning","Niveau de carburant faible","fuel_system",
     "medium","Ravitaillement",
     "Prevoir un ravitaillement rapidement pour eviter l'arret du vehicule.",
     "sensor_rules", None),

    ("speed_critical", "speed", None, 150, "gt_max",
     "driving_speed","critical","Vitesse vehicule excessivement elevee","",
     "high","Conduite a risque",
     "Reduire immediatement la vitesse et controler les conditions de roulage.",
     "sensor_rules", None),

    ("speed_warning",  "speed", None, 120, "gt_max",
     "driving_speed","warning","Vitesse vehicule elevee","",
     "medium","Conduite",
     "Adapter la vitesse pour limiter l'usure et les risques mecaniques.",
     "sensor_rules", None),

    ("rpm_warning",    "rpm", None, 4500, "gt_max",
     "driving_style","warning","Regime moteur eleve","",
     "medium","Style de conduite",
     "Analyser le style de conduite et eviter les sur-regimes prolonges.",
     "sensor_rules", None),

    ("load_warning",   "engine_load", None, None, None,
     "engine_load","warning","Charge moteur elevee","engine_system",
     "medium","Charge moteur",
     "Reduire la charge ou verifier les conditions de roulage et le systeme moteur.",
     "sensor_rules",
     cj({"field":"engine_load","op":"gte","value":85})),

    ("cpu_temp_critical","temp_cpu", None, 90, "gt_max",
     "device_cpu_temp","critical","Temperature CPU boitier critique","telematics_device",
     "high","Boitier telematique",
     "Controler le refroidissement du boitier et l'exposition a la chaleur.",
     "sensor_rules", None),

    ("cpu_temp_warning", "temp_cpu", None, 80, "gt_max",
     "device_cpu_temp","warning","Temperature CPU boitier elevee","telematics_device",
     "medium","Boitier telematique",
     "Surveiller la temperature CPU du boitier.",
     "sensor_rules", None),

    ("cpu_load_warning", "cpu", None, 90, "gt_max",
     "device_cpu_load","warning","Charge CPU boitier elevee","",
     "medium","Charge systeme",
     "Verifier les taches telematiques actives (CPU eleve).",
     "sensor_rules", None),

    ("gpu_load_warning", "gpu", None, 90, "gt_max",
     "device_gpu_load","warning","Charge GPU boitier elevee","telematics_device",
     "medium","Charge systeme",
     "Verifier les charges GPU anormales sur le boitier.",
     "sensor_rules", None),

    ("intake_critical",  "intake_temp", None, 75, "gt_max",
     "intake","critical","Temperature d'admission critique","intake_system",
     "high","Admission d'air",
     "Controler le circuit d'admission et la circulation d'air moteur.",
     "sensor_rules", None),

    ("intake_warning",   "intake_temp", None, 60, "gt_max",
     "intake","warning","Temperature d'admission elevee","intake_system",
     "medium","Admission d'air",
     "Inspecter le filtre et la prise d'air.",
     "sensor_rules", None),

    ("ambient_warning",  "ambient_air_temp", None, 40, "gt_max",
     "ambient_heat","warning","Temperature ambiante elevee","",
     "low","Conditions externes",
     "Adapter la conduite et surveiller les temperatures moteur en periode chaude.",
     "sensor_rules", None),

    ("thermal_delta_warning","thermal_delta", None, None, None,
     "thermal_delta","warning","Ecart thermique moteur/admission eleve","cooling_system",
     "medium","Diagnostic thermique",
     "Verifier capteurs temperature moteur/admission et circuit de refroidissement.",
     "sensor_rules",
     cj({"all":[{"field":"thermal_delta","op":"gt","value":45},
                {"any":[{"field":"engine_temp","op":"gte","value":99},
                        {"field":"intake_temp","op":"gte","value":60}]}]})),

    # ── maintenance_rules (7) ─────────────────────────────────────────────────
    ("oil_critical",  "oil_ratio", None, None, None,
     "maintenance_oil","critical","Vidange tres en retard","oil_service",
     "high","Vidange urgente",
     "Depasser fortement l'intervalle de vidange. Faire la vidange immediatement.",
     "maintenance_rules",
     cj({"field":"oil_ratio","op":"gte","value":1.2})),

    ("oil_warning",   "oil_ratio", None, None, None,
     "maintenance_oil","warning","Vidange due","oil_service",
     "medium","Vidange",
     "Intervalle de vidange atteint. Planifier la vidange rapidement.",
     "maintenance_rules",
     cj({"field":"oil_ratio","op":"gte","value":1.0})),

    ("oil_info",      "oil_ratio", None, None, None,
     "maintenance_oil","info","Vidange bientot due","oil_service",
     "low","Preparation vidange",
     "La vidange approche. Preparer le rendez-vous d'entretien.",
     "maintenance_rules",
     cj({"field":"oil_ratio","op":"gte","value":0.85})),

    ("service_warning","service_ratio", None, None, None,
     "maintenance_general","warning","Entretien general en retard","general_service",
     "medium","Entretien general",
     "L'intervalle d'entretien general est depasse. Planifier un controle complet.",
     "maintenance_rules",
     cj({"field":"service_ratio","op":"gte","value":1.15})),

    ("service_info",  "service_ratio", None, None, None,
     "maintenance_general","info","Entretien general bientot du","general_service",
     "low","Preparation entretien",
     "Prevoir le prochain entretien general.",
     "maintenance_rules",
     cj({"field":"service_ratio","op":"gte","value":0.9})),

    ("parts_warning", "parts_ratio", None, None, None,
     "maintenance_parts","warning","Controle des pieces majeures recommande","major_parts",
     "medium","Pieces majeures",
     "Verifier l'etat des pieces majeures selon l'intervalle configure.",
     "maintenance_rules",
     cj({"field":"parts_ratio","op":"gte","value":1.1})),

    ("parts_info",    "parts_ratio", None, None, None,
     "maintenance_parts","info","Pieces majeures bientot a controler","major_parts",
     "low","Suivi pieces",
     "Prevoir un controle preventif des pieces majeures.",
     "maintenance_rules",
     cj({"field":"parts_ratio","op":"gte","value":0.9})),

    # ── fallback_rules (2) ────────────────────────────────────────────────────
    ("mileage_warning","odometer", None, None, None,
     "mileage","warning","Kilometrage tres eleve - maintenance lourde a planifier","general_service",
     "medium","Maintenance kilometrage",
     "Prevoir un controle complet des organes moteur, refroidissement, admission et charge.",
     "fallback_rules",
     cj({"all":[{"field":"has_maintenance_baseline","op":"eq","value":False},
                {"field":"odometer","op":"gte","value":200000}]})),

    ("mileage_info",  "odometer", None, None, None,
     "mileage","info","Kilometrage eleve - maintenance preventive conseillee","general_service",
     "low","Entretien preventif",
     "Verifier l'entretien periodique du vehicule en fonction du kilometrage.",
     "fallback_rules",
     cj({"all":[{"field":"has_maintenance_baseline","op":"eq","value":False},
                {"field":"odometer","op":"gte","value":120000}]})),
]

assert len(ROWS) == 30, f"Expected 30, got {len(ROWS)}"

# ── Write recommendation_rules_simple (engine reads this) ────────────────────
print(f"Patching: {XLSX.name}")
wb = load_workbook(XLSX)

name = "recommendation_rules_simple"
if name in wb.sheetnames:
    del wb[name]
ws = wb.create_sheet(name)
ws.append(COLS)
for row in ROWS:
    ws.append(list(row))

cfg_name = "engine_config"
if cfg_name in wb.sheetnames:
    del wb[cfg_name]
ws_cfg = wb.create_sheet(cfg_name)
ws_cfg.append(ENGINE_CONFIG_COLS)
for row in ENGINE_CONFIG_ROWS:
    ws_cfg.append(list(row))

wb.save(XLSX)

# ── Verify via engine ─────────────────────────────────────────────────────────
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from app.services.ia import recommendation_engine as eng
eng._load_rules.cache_clear()
rules = eng._load_rules()

s = len(rules.get("sensor_rules", []))
m = len(rules.get("maintenance_rules", []))
f = len(rules.get("fallback_rules", []))
total = s + m + f

print(f"\nRESULT after patch:")
print(f"  sensor_rules     : {s}")
print(f"  maintenance_rules: {m}  {[r['id'] for r in rules.get('maintenance_rules',[])]}")
print(f"  fallback_rules   : {f}  {[r['id'] for r in rules.get('fallback_rules',[])]}")
print(f"  TOTAL            : {total}/30")
if total == 30:
    print("\nOK - all 30 rules loaded by engine from dataset")
else:
    missing = 30 - total
    print(f"\nWARNING - {missing} rules missing")

