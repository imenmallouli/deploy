"""
Generate reference Excel file with 3 sheets:
  - Rules          : all 30 rules (human-readable format)
  - Sample_Dataset : 10 sample telemetry rows
  - Thresholds     : parameter thresholds

Color scheme (matching screenshots):
  critical → light pink   #FFD7D7
  warning  → light yellow #FFF9C4
  info     → light green  #E8F5E9
"""

from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = Path(__file__).resolve().parent.parent / "data" / "reference_rules_dataset.xlsx"

# ─── colour palette ────────────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="1A2744")   # dark navy
HDR_FONT   = Font(color="FFFFFF", bold=True, size=10)
CRIT_FILL  = PatternFill("solid", fgColor="FFD7D7")   # light pink
WARN_FILL  = PatternFill("solid", fgColor="FFF9C4")   # light yellow
INFO_FILL  = PatternFill("solid", fgColor="E8F5E9")   # light green
CRIT_FONT  = Font(color="C0392B", bold=True, size=10)
WARN_FONT  = Font(color="B7770D", bold=True, size=10)
INFO_FONT  = Font(color="1E7E34", bold=True, size=10)
NORM_FONT  = Font(size=10)
THIN       = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

def _hdr(ws, row, cols):
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN

def _row(ws, row_idx, values, severity=None):
    sev = (severity or "").lower()
    row_fill = CRIT_FILL if sev == "critical" else (WARN_FILL if sev == "warning" else (INFO_FILL if sev == "info" else None))
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=c, value=val)
        if row_fill:
            cell.fill = row_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.font = NORM_FONT
        cell.border = THIN
        # severity column gets coloured font (col C in Rules)
        if c == 3 and sev:
            cell.font = CRIT_FONT if sev == "critical" else (WARN_FONT if sev == "warning" else INFO_FONT)

def _set_col_width(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _freeze(ws, cell="B2"):
    ws.freeze_panes = cell

# ─── ALL 30 RULES ──────────────────────────────────────────────────────────────
RULES = [
    # Category, Condition, Severity, Message, Value_Field, Component, Sug_Priority, Sug_Title, Sug_Detail
    # ── sensor_rules ──────────────────────────────────────────────────────────
    ("Battery",            "battery < 11.5",                             "critical", "Batterie critique detectee",
     "battery", "battery_system", "high", "Batterie / alternateur",
     "Controler immediatement la batterie et le circuit de charge."),

    ("Battery",            "battery < 12.0",                             "warning",  "Batterie en baisse detectee",
     "battery", "battery_system", "medium", "Controle batterie",
     "Planifier une verification batterie dans les prochaines 48h."),

    ("Battery",            "12.0 <= battery < 12.5 AND rpm > 500",       "warning",  "Tension basse moteur allume - alternateur suspect",
     "battery", "", "medium", "Test alternateur",
     "Faire diagnostiquer l'alternateur ; la batterie ne se recharge pas correctement."),

    ("Battery",            "battery > 15.8",                             "critical", "Surtension batterie critique detectee",
     "battery", "", "high", "Systeme electrique",
     "Verifier l'alternateur et le systeme de charge. Risque de dommage electrique."),

    ("Battery",            "battery > 15.5",                             "warning",  "Surtension batterie detectee",
     "battery", "", "medium", "Test alternateur",
     "Faire verifier le systeme de charge et l'alternateur rapidement."),

    ("Engine Temp",        "temp > 106",                                  "critical", "Surchauffe moteur detectee",
     "temp", "cooling_system", "high", "Systeme de refroidissement",
     "Verifier liquide de refroidissement, ventilateur et radiateur."),

    ("Engine Temp",        "temp >= 99",                                  "warning",  "Temperature moteur elevee",
     "temp", "cooling_system", "medium", "Inspection moteur",
     "Prevoir une inspection du systeme de refroidissement."),

    ("Fuel",               "fuel < 5",                                    "critical", "Niveau de carburant critique",
     "fuel", "fuel_system", "high", "Ravitaillement urgent",
     "Ravitailler immediatement - risque d'arret moteur."),

    ("Fuel",               "fuel < 12",                                   "warning",  "Niveau de carburant faible",
     "fuel", "fuel_system", "medium", "Ravitaillement",
     "Prevoir un ravitaillement rapidement pour eviter l'arret du vehicule."),

    ("Driving Speed",      "speed > 150",                                 "critical", "Vitesse vehicule excessivement elevee",
     "speed", "", "high", "Conduite a risque",
     "Reduire immediatement la vitesse et controler les conditions de roulage."),

    ("Driving Speed",      "speed > 120",                                 "warning",  "Vitesse vehicule elevee",
     "speed", "", "medium", "Conduite",
     "Adapter la vitesse pour limiter l'usure et les risques mecaniques."),

    ("RPM",                "rpm > 4500",                                  "warning",  "Regime moteur eleve",
     "rpm", "", "medium", "Style de conduite",
     "Analyser le style de conduite et eviter les sur-regimes prolonges."),

    ("Engine Load",        "engine_load > 85",                            "warning",  "Charge moteur elevee",
     "engine_load", "engine_system", "medium", "Charge moteur",
     "Reduire la charge ou verifier les conditions de roulage et le systeme moteur."),

    ("Device Temp",        "temp_cpu > 90",                               "critical", "Temperature CPU boitier critique",
     "temp_cpu", "telematics_device", "high", "Boitier telematique",
     "Controler le refroidissement du boitier et l'exposition a la chaleur."),

    ("Device Temp",        "temp_cpu > 80",                               "warning",  "Temperature CPU boitier elevee",
     "temp_cpu", "telematics_device", "medium", "Boitier telematique",
     "Surveiller la temperature CPU du boitier."),

    ("Device Load",        "cpu > 90",                                    "warning",  "Charge CPU boitier elevee",
     "cpu", "", "medium", "Charge systeme",
     "Verifier les taches telematiques actives (CPU eleve)."),

    ("Device Load",        "gpu > 90",                                    "warning",  "Charge GPU boitier elevee",
     "gpu", "telematics_device", "medium", "Charge systeme",
     "Verifier les charges GPU anormales sur le boitier."),

    ("Intake Temp",        "intake_temp > 75",                            "critical", "Temperature d'admission critique",
     "intake_temp", "intake_system", "high", "Admission d'air",
     "Controler le circuit d'admission et la circulation d'air moteur."),

    ("Intake Temp",        "intake_temp > 60",                            "warning",  "Temperature d'admission elevee",
     "intake_temp", "intake_system", "medium", "Admission d'air",
     "Inspecter le filtre et la prise d'air."),

    ("Ambient Temp",       "ambient_air_temp > 40",                       "warning",  "Temperature ambiante elevee",
     "ambient_air_temp", "", "low", "Conditions externes",
     "Adapter la conduite et surveiller les temperatures moteur en periode chaude."),

    ("Thermal Delta",      "thermal_delta > 45 AND (engine_temp >= 99 OR intake_temp >= 60)", "warning", "Ecart thermique moteur/admission eleve",
     "thermal_delta", "cooling_system", "medium", "Diagnostic thermique",
     "Verifier capteurs temperature moteur/admission et circuit de refroidissement."),

    # ── maintenance_rules ─────────────────────────────────────────────────────
    ("Maintenance (Oil)",     "oil_ratio > 120%  (odometer - last_oil > 1.2 x interval)", "critical", "Vidange tres en retard",
     "oil_ratio", "oil_service", "high", "Vidange urgente",
     "Intervalle de vidange fortement depasse. Faire la vidange immediatement."),

    ("Maintenance (Oil)",     "oil_ratio > 100%  (odometer - last_oil > interval)",        "warning",  "Vidange due",
     "oil_ratio", "oil_service", "medium", "Vidange",
     "Intervalle de vidange atteint. Planifier la vidange rapidement."),

    ("Maintenance (Oil)",     "oil_ratio > 85%   (odometer - last_oil > 0.85 x interval)", "info",     "Vidange bientot due",
     "oil_ratio", "oil_service", "low", "Preparation vidange",
     "La vidange approche. Preparer le rendez-vous d'entretien."),

    ("Maintenance (Service)", "service_ratio > 115%  (odometer - last_service > 1.15 x interval)", "warning", "Entretien general en retard",
     "service_ratio", "general_service", "medium", "Entretien general",
     "L'intervalle d'entretien general est depasse. Planifier un controle complet."),

    ("Maintenance (Service)", "service_ratio > 90%   (odometer - last_service > 0.9 x interval)",  "info",    "Entretien general bientot du",
     "service_ratio", "general_service", "low", "Preparation entretien",
     "Prevoir le prochain entretien general."),

    ("Maintenance (Parts)",   "parts_ratio > 110%  (odometer - last_parts > 1.1 x interval)", "warning", "Controle des pieces majeures recommande",
     "parts_ratio", "major_parts", "medium", "Pieces majeures",
     "Verifier l'etat des pieces majeures selon l'intervalle configure."),

    ("Maintenance (Parts)",   "parts_ratio > 90%   (odometer - last_parts > 0.9 x interval)", "info",    "Pieces majeures bientot a controler",
     "parts_ratio", "major_parts", "low", "Suivi pieces",
     "Prevoir un controle preventif des pieces majeures."),

    # ── fallback_rules ────────────────────────────────────────────────────────
    ("Mileage",            "odometer > 200000  (no maintenance baseline)", "warning", "Kilometrage tres eleve - maintenance lourde a planifier",
     "odometer", "general_service", "medium", "Maintenance kilometrage",
     "Prevoir un controle complet des organes moteur, refroidissement, admission et charge."),

    ("Mileage",            "odometer > 120000  (no maintenance baseline)", "info",    "Kilometrage eleve - maintenance preventive conseillee",
     "odometer", "general_service", "low", "Entretien preventif",
     "Verifier l'entretien periodique du vehicule en fonction du kilometrage."),
]

# ─── SAMPLE TELEMETRY ──────────────────────────────────────────────────────────
TELEMETRY_COLS = [
    "vehicle_id", "timestamp", "battery", "temp", "fuel", "speed",
    "rpm", "temp_cpu", "cpu", "gpu", "intake_temp", "ambient_air_temp",
    "odometer", "last_oil",
]
TELEMETRY_ROWS = [
    ("VH001", "2024-01-15 08:00", 11.2, 107,  4,  160, 4600, 92, 95, 91, 78, 42, 215000, 210000),
    ("VH002", "2024-01-15 09:30", 11.8, 100, 10,  125, 3200, 85, 75, 60, 62, 38, 135000, 132000),
    ("VH003", "2024-01-15 10:00", 12.4,  95, 20,   90, 2500, 70, 50, 40, 55, 30,  80000,  75000),
    ("VH004", "2024-01-15 11:00", 16.0,  85, 35,   70, 2000, 65, 45, 35, 50, 25,  45000,  40000),
    ("VH005", "2024-01-15 12:00", 15.6,  80, 50,   60, 1800, 60, 40, 30, 45, 20,  25000,  22000),
    ("VH006", "2024-01-15 13:00", 12.8,  98,  3,  140, 4700, 88, 92, 50, 58, 35, 180000, 168000),
    ("VH007", "2024-01-15 14:00", 12.1, 102,  8,   55, 1500, 55, 35, 25, 40, 22,  92000,  91000),
    ("VH008", "2024-01-15 15:00", 12.6,  93, 45,  100, 3800, 78, 60, 55, 65, 41, 160000, 148000),
    ("VH009", "2024-01-15 16:00", 11.5,  88, 60,   80, 2200, 72, 48, 38, 52, 28,  55000,  46000),
    ("VH010", "2024-01-15 17:00", 13.0,  75, 70,   50, 1600, 50, 30, 20, 38, 18,  30000,  29500),
]

# ─── THRESHOLDS ───────────────────────────────────────────────────────────────
THRESH_COLS = ["Parameter", "Field_Name", "Critical_Low", "Warning_Low", "Warning_High", "Critical_High", "Unit", "Notes"]
THRESH_ROWS = [
    ("Battery Voltage",      "battery",          11.5, 12,    15.5,  15.8,  "V",       "Also check: 12.0<=battery<12.5 AND rpm>500 → alternateur suspect"),
    ("Engine Temperature",   "temp",             None, 99,    None,  106,   "°C",      ""),
    ("Fuel Level",           "fuel",             5,    12,    None,  None,  "L or %",  ""),
    ("Vehicle Speed",        "speed",            None, None,  120,   150,   "km/h",    ""),
    ("RPM",                  "rpm",              None, None,  4500,  None,  "RPM",     ""),
    ("CPU Temp (Device)",    "temp_cpu",         None, None,  80,    90,    "°C",      ""),
    ("CPU Load (Device)",    "cpu",              None, None,  None,  90,    "%",       ""),
    ("GPU Load (Device)",    "gpu",              None, None,  None,  90,    "%",       ""),
    ("Intake Temperature",   "intake_temp",      None, None,  60,    75,    "°C",      ""),
    ("Ambient Air Temp",     "ambient_air_temp", None, None,  40,    None,  "°C",      ""),
    ("Oil Interval",         "oil_delta",        None, "85% of interval",  "100% of interval", "120% of interval", "km", "oil_delta = odometer - last_oil"),
    ("Service Interval",     "service_delta",    None, "90% of interval",  None,       "115% of interval", "km", "service_delta = odometer - last_service"),
    ("Odometer (no baseline)","odometer",        None, None,  120000, 200000, "km",    "Only when no maintenance baseline"),
]


def build_rules_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Rules")
    cols = ["Category", "Condition", "Severity", "Message",
            "Value_Field", "Component", "Suggestion_Priority",
            "Suggestion_Title", "Suggestion_Detail"]
    _hdr(ws, 1, cols)
    ws.row_dimensions[1].height = 30

    for i, rule in enumerate(RULES, 2):
        sev = rule[2]
        _row(ws, i, list(rule), severity=sev)
        ws.row_dimensions[i].height = 40

    _set_col_width(ws, [18, 38, 10, 36, 14, 18, 14, 22, 55])
    _freeze(ws, "A2")
    ws.sheet_view.showGridLines = False


def build_sample_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Sample_Dataset")
    _hdr(ws, 1, TELEMETRY_COLS)
    ws.row_dimensions[1].height = 25

    for i, row in enumerate(TELEMETRY_ROWS, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=c, value=val)
            cell.alignment = Alignment(vertical="center", horizontal="center")
            cell.font = NORM_FONT
            cell.border = THIN
        ws.row_dimensions[i].height = 20

    _set_col_width(ws, [8, 18, 9, 6, 6, 7, 7, 9, 5, 5, 12, 16, 10, 10])
    _freeze(ws, "B2")
    ws.sheet_view.showGridLines = False


def build_thresholds_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Thresholds")
    _hdr(ws, 1, THRESH_COLS)
    ws.row_dimensions[1].height = 30

    # Header row colours for crit/warn columns
    crit_col_fill = PatternFill("solid", fgColor="FFD7D7")
    warn_col_fill = PatternFill("solid", fgColor="FFF9C4")
    # col C & F = critical, col D & E = warning
    for col, fill in [(3, crit_col_fill), (4, warn_col_fill), (5, warn_col_fill), (6, crit_col_fill)]:
        ws.cell(1, col).fill = fill
        ws.cell(1, col).font = Font(bold=True, size=10)

    for i, row in enumerate(THRESH_ROWS, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=c, value=val)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.font = NORM_FONT
            cell.border = THIN
            if c == 3 and val not in (None, ""):
                cell.fill = crit_col_fill
            elif c in (4, 5) and val not in (None, ""):
                cell.fill = warn_col_fill
            elif c == 6 and val not in (None, ""):
                cell.fill = crit_col_fill
        ws.row_dimensions[i].height = 22

    _set_col_width(ws, [22, 18, 13, 18, 18, 18, 9, 45])
    _freeze(ws, "A2")
    ws.sheet_view.showGridLines = False


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)           # remove default sheet

    build_rules_sheet(wb)
    build_sample_sheet(wb)
    build_thresholds_sheet(wb)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"OK  → {OUTPUT}")
    print(f"     Rules          : {len(RULES)} rows")
    print(f"     Sample_Dataset : {len(TELEMETRY_ROWS)} rows")
    print(f"     Thresholds     : {len(THRESH_ROWS)} rows")


if __name__ == "__main__":
    main()
