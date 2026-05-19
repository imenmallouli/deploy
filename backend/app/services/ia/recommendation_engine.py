from __future__ import annotations

import json
from functools import lru_cache
from pathlib import Path


DATA_DIR = Path(__file__).resolve().parents[3] / "data"


def _find_latest_dataset_xlsx() -> Path | None:
    candidates = []
    for path in DATA_DIR.glob("sample_dataset*.xlsx"):
        if path.name.startswith("~$"):
            continue
        candidates.append(path)
    if not candidates:
        return None
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0]


def _load_rules_from_dataset(dataset_path: Path) -> dict | None:
    try:
        from openpyxl import load_workbook
    except Exception:
        return None

    try:
        wb = load_workbook(dataset_path, read_only=True, data_only=True)
    except Exception:
        return None

    def _as_bool(value) -> bool:
        if isinstance(value, bool):
            return value
        text = str(value or "").strip().lower()
        return text in {"1", "true", "yes", "y", "oui"}

    def _as_float_or_none(value):
        if value is None or str(value).strip() == "":
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    def _new_rules_data() -> dict:
        return {
            "sensor_rules": [],
            "maintenance_rules": [],
            "fallback_rules": [],
            "score_floor": {
                "critical_count_ge_2": 85.0,
                "critical_any": 70.0,
                "warning_any": 40.0,
            },
            "severity_promotion": {
                "critical_min_score": 70.0,
                "warning_min_score": 40.0,
            },
            "defaults": {
                "normal_suggestion": {
                    "priority": "low",
                    "title": "Etat normal",
                    "message": "Aucune anomalie detectee actuellement.",
                }
            },
        }

    def _set_nested(target: dict, path_parts: list[str], value):
        cursor = target
        for part in path_parts[:-1]:
            if part not in cursor or not isinstance(cursor.get(part), dict):
                cursor[part] = {}
            cursor = cursor[part]
        cursor[path_parts[-1]] = value

    def _apply_engine_config(data: dict):
        if "engine_config" not in wb.sheetnames:
            return

        ws_cfg = wb["engine_config"]
        cfg_rows = list(ws_cfg.iter_rows(values_only=True))
        if not cfg_rows:
            return

        headers = [str(h or "").strip().lower() for h in cfg_rows[0]]
        index = {name: i for i, name in enumerate(headers) if name}

        def _cfg_get(row, key, default=None):
            idx = index.get(key)
            if idx is None or idx >= len(row):
                return default
            value = row[idx]
            return default if value is None else value

        for row in cfg_rows[1:]:
            key = str(_cfg_get(row, "key", "")).strip()
            if not key:
                continue

            # Only allow known config trees from dataset.
            if not key.startswith(("score_floor.", "severity_promotion.", "defaults.")):
                continue

            value = _cfg_get(row, "value_json", None)
            if value is None:
                value = _cfg_get(row, "value", None)

            parsed_value = value
            if isinstance(value, str):
                raw = value.strip()
                if raw:
                    try:
                        parsed_value = json.loads(raw)
                    except (json.JSONDecodeError, ValueError):
                        parsed_value = raw

            parts = [p for p in key.split(".") if p]
            if not parts:
                continue
            _set_nested(data, parts, parsed_value)

    try:
        # Preferred format 1: very simple table requested by user
        # (variable, seuil_min, seuil_max, type_risque, message).
        if "recommendation_rules_simple" in wb.sheetnames:
            ws = wb["recommendation_rules_simple"]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                headers = [str(h or "").strip().lower() for h in rows[0]]
                index = {name: i for i, name in enumerate(headers) if name}

                def _get(row, key, default=None):
                    idx = index.get(key)
                    if idx is None or idx >= len(row):
                        return default
                    value = row[idx]
                    return default if value is None else value

                data = _new_rules_data()

                for i, row in enumerate(rows[1:], start=2):
                    variable = str(_get(row, "variable", "")).strip()
                    if not variable:
                        continue

                    seuil_min = _as_float_or_none(_get(row, "seuil_min", None))
                    seuil_max = _as_float_or_none(_get(row, "seuil_max", None))
                    mode = str(_get(row, "mode", "")).strip().lower()

                    if mode == "inside_range" and seuil_min is not None and seuil_max is not None:
                        when = {
                            "all": [
                                {"field": variable, "op": "gte", "value": seuil_min},
                                {"field": variable, "op": "lte", "value": seuil_max},
                            ]
                        }
                    elif mode == "outside_range" and seuil_min is not None and seuil_max is not None:
                        when = {
                            "any": [
                                {"field": variable, "op": "lt", "value": seuil_min},
                                {"field": variable, "op": "gt", "value": seuil_max},
                            ]
                        }
                    elif mode == "lt_min" and seuil_min is not None:
                        when = {"field": variable, "op": "lt", "value": seuil_min}
                    elif mode == "gt_max" and seuil_max is not None:
                        when = {"field": variable, "op": "gt", "value": seuil_max}
                    elif seuil_max is not None and seuil_min is None:
                        # Default interpretation used in user example: value > seuil_max.
                        when = {"field": variable, "op": "gt", "value": seuil_max}
                    elif seuil_min is not None and seuil_max is None:
                        # Default interpretation: value < seuil_min.
                        when = {"field": variable, "op": "lt", "value": seuil_min}
                    elif seuil_min is not None and seuil_max is not None:
                        # With both thresholds and no mode, use outside range.
                        when = {
                            "any": [
                                {"field": variable, "op": "lt", "value": seuil_min},
                                {"field": variable, "op": "gt", "value": seuil_max},
                            ]
                        }
                    else:
                        when = None

                    # condition_json overrides (or provides) the when dict.
                    cj_raw = str(_get(row, "condition_json", "")).strip()
                    if cj_raw:
                        try:
                            when = json.loads(cj_raw)
                        except (json.JSONDecodeError, ValueError):
                            pass

                    if when is None:
                        continue

                    risk_type = str(_get(row, "type_risque", "unknown")).strip()
                    risk_severity = str(_get(row, "niveau", "warning")).strip() or "warning"
                    message = str(_get(row, "message", "Alerte detectee")).strip() or "Alerte detectee"
                    component = str(_get(row, "component", "")).strip()
                    suggestion_priority = str(_get(row, "suggestion_priority", "")).strip()
                    suggestion_title = str(_get(row, "suggestion_title", "")).strip()
                    suggestion_message = str(_get(row, "suggestion_message", "")).strip()

                    rule = {
                        "id": str(_get(row, "rule_id", f"row_{i}")),
                        "when": when,
                        "risk": {
                            "type": risk_type,
                            "severity": risk_severity,
                            "message": message,
                        },
                        "value_field": variable,
                    }
                    if component:
                        rule["component"] = component
                    if suggestion_priority or suggestion_title or suggestion_message:
                        rule["suggestion"] = {
                            "priority": suggestion_priority or "low",
                            "title": suggestion_title or "Action recommandee",
                            "message": suggestion_message or "Verifier le vehicule.",
                        }

                    ruleset = str(_get(row, "ruleset", "sensor_rules")).strip()
                    if ruleset not in {"sensor_rules", "maintenance_rules", "fallback_rules"}:
                        ruleset = "sensor_rules"
                    data[ruleset].append(rule)

                _apply_engine_config(data)
                if data["sensor_rules"] or data["maintenance_rules"] or data["fallback_rules"]:
                    return data

        # Preferred format 2: one rule per row (advanced table).
        if "recommendation_rules_table" in wb.sheetnames:
            ws = wb["recommendation_rules_table"]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                headers = [str(h or "").strip() for h in rows[0]]
                index = {name: i for i, name in enumerate(headers) if name}

                def _get(row, key, default=None):
                    idx = index.get(key)
                    if idx is None or idx >= len(row):
                        return default
                    value = row[idx]
                    return default if value is None else value

                data = _new_rules_data()

                for row in rows[1:]:
                    rule_id = str(_get(row, "rule_id", "")).strip()
                    if not rule_id:
                        continue

                    ruleset = str(_get(row, "ruleset", "sensor_rules")).strip()
                    if ruleset not in {"sensor_rules", "maintenance_rules", "fallback_rules"}:
                        ruleset = "sensor_rules"

                    condition_json = str(_get(row, "condition_json", "")).strip()
                    if condition_json:
                        when = json.loads(condition_json)
                    else:
                        field = str(_get(row, "field", "")).strip()
                        op = str(_get(row, "op", "eq")).strip().lower()
                        value = _as_float_or_none(_get(row, "value", None))
                        if value is None:
                            raw_value = _get(row, "value", None)
                            if isinstance(raw_value, str) and raw_value.strip().lower() in {"true", "false"}:
                                value = raw_value.strip().lower() == "true"
                            else:
                                value = raw_value
                        when = {"field": field, "op": op, "value": value}

                    rule = {
                        "id": rule_id,
                        "when": when,
                        "risk": {
                            "type": str(_get(row, "risk_type", "unknown")),
                            "severity": str(_get(row, "risk_severity", "warning")),
                            "message": str(_get(row, "message", "Alerte detectee")),
                        },
                    }

                    group = str(_get(row, "group", "")).strip()
                    if group:
                        rule["group"] = group

                    if _as_bool(_get(row, "exclusive_group", False)):
                        rule["exclusive_group"] = True

                    component = str(_get(row, "component", "")).strip()
                    if component:
                        rule["component"] = component

                    value_field = str(_get(row, "value_field", "")).strip()
                    if value_field:
                        rule["value_field"] = value_field

                    sug_priority = str(_get(row, "suggestion_priority", "")).strip()
                    sug_title = str(_get(row, "suggestion_title", "")).strip()
                    sug_message = str(_get(row, "suggestion_message", "")).strip()
                    if sug_priority or sug_title or sug_message:
                        rule["suggestion"] = {
                            "priority": sug_priority or "low",
                            "title": sug_title or "Action recommandee",
                            "message": sug_message or "Verifier le vehicule.",
                        }

                    data[ruleset].append(rule)

                _apply_engine_config(data)
                if data["sensor_rules"] or data["maintenance_rules"] or data["fallback_rules"]:
                    return data

        # Backward compatible format: full JSON in A2.
        if "recommendation_rules" in wb.sheetnames:
            ws = wb["recommendation_rules"]
            raw = ws["A2"].value
            if not raw:
                return None
            return json.loads(str(raw))
        return None
    except Exception:
        return None
    finally:
        wb.close()


@lru_cache(maxsize=1)
def _load_rules() -> dict:
    default_rules = {
        "sensor_rules": [],
        "maintenance_rules": [],
        "fallback_rules": [],
        "score_floor": {
            "critical_count_ge_2": 85.0,
            "critical_any": 70.0,
            "warning_any": 40.0,
        },
        "severity_promotion": {
            "critical_min_score": 70.0,
            "warning_min_score": 40.0,
        },
        "defaults": {
            "normal_suggestion": {
                "priority": "low",
                "title": "Etat normal",
                "message": "Aucune anomalie detectee actuellement.",
            }
        },
    }

    dataset_path = _find_latest_dataset_xlsx()
    if dataset_path is None:
        return default_rules

    dataset_rules = _load_rules_from_dataset(dataset_path)
    if not isinstance(dataset_rules, dict) or not dataset_rules:
        return default_rules
    return dataset_rules


class RecommendationEngine:
    @staticmethod
    def _matches_condition(condition: dict, values: dict) -> bool:
        if not condition:
            return True

        if "all" in condition:
            return all(RecommendationEngine._matches_condition(c, values) for c in condition["all"])
        if "any" in condition:
            return any(RecommendationEngine._matches_condition(c, values) for c in condition["any"])

        field = condition.get("field")
        op = str(condition.get("op") or "eq").lower()
        target = condition.get("value")
        left = values.get(field)

        if left is None:
            return False

        if op == "lt":
            return left < target
        if op == "lte":
            return left <= target
        if op == "gt":
            return left > target
        if op == "gte":
            return left >= target
        if op == "eq":
            return left == target
        if op == "neq":
            return left != target
        return False

    @staticmethod
    def enrich_prediction(prediction: dict) -> dict:
        rules = _load_rules()

        snapshot = prediction.get("telemetry_snapshot", {})
        severity = str(prediction.get("predicted_severity", "info"))
        risk_score = float(prediction.get("predicted_risk_score", 0))

        battery = snapshot.get("battery_voltage")
        temp = snapshot.get("engine_temp")
        fuel = snapshot.get("fuel_level")
        speed = snapshot.get("speed")
        rpm = snapshot.get("rpm")
        load = snapshot.get("engine_load")
        intake_temp = snapshot.get("intake_temp")
        ambient_air_temp = snapshot.get("ambient_air_temp")
        odometer = snapshot.get("odometer")
        temp_cpu = snapshot.get("temp_cpu")
        cpu = snapshot.get("cpu")
        gpu = snapshot.get("gpu")
        active_dtc_events = prediction.get("active_dtc_events") or []
        maintenance_context = prediction.get("maintenance_context") or {}
        maintenance_records = prediction.get("maintenance_records") or []

        predicted_risks: list[dict] = []
        maintenance_suggestions: list[dict] = []
        suppressed_alerts: list[dict] = []

        def _is_recently_serviced(component: str | None = None, dtc_code: str | None = None) -> tuple[bool, str | None]:
            if not maintenance_records:
                return False, None
            if odometer is None:
                return False, None

            target_component = (component or "").strip().lower()
            target_code = (dtc_code or "").strip().upper()

            for record in maintenance_records:
                if not isinstance(record, dict):
                    continue

                record_component = str(record.get("component") or "").strip().lower()
                serviced_at = record.get("serviced_at_odometer")
                valid_for_km = record.get("valid_for_km")
                resolved_codes = [str(c).strip().upper() for c in (record.get("resolved_dtc_codes") or []) if str(c).strip()]

                try:
                    serviced_at_val = float(serviced_at) if serviced_at is not None else None
                    valid_for_km_val = float(valid_for_km) if valid_for_km is not None else 0.0
                except (TypeError, ValueError):
                    continue

                if serviced_at_val is None or valid_for_km_val <= 0:
                    continue

                delta = float(odometer) - serviced_at_val
                if delta < 0 or delta > valid_for_km_val:
                    continue

                component_match = bool(target_component) and record_component in {target_component, "all"}
                code_match = bool(target_code) and target_code in resolved_codes

                if component_match or code_match:
                    reason = f"recent_service:{record_component or 'unknown'} ({int(delta)}km/{int(valid_for_km_val)}km)"
                    return True, reason

            return False, None

        def add_risk(risk_type: str, risk_severity: str, message: str, value=None, component: str | None = None, dtc_code: str | None = None):
            suppressed, reason = _is_recently_serviced(component=component, dtc_code=dtc_code)
            if suppressed:
                suppressed_alerts.append(
                    {
                        "risk_type": risk_type,
                        "message": message,
                        "component": component,
                        "dtc_code": dtc_code,
                        "reason": reason,
                    }
                )
                return False

            predicted_risks.append(
                {
                    "type": risk_type,
                    "severity": risk_severity,
                    "message": message,
                    "value": value,
                }
            )
            return True

        def add_suggestion(priority: str, title: str, message: str):
            maintenance_suggestions.append(
                {
                    "priority": priority,
                    "title": title,
                    "message": message,
                }
            )

        values = {
            "battery_voltage": battery,
            "engine_temp": temp,
            "fuel_level": fuel,
            "speed": speed,
            "rpm": rpm,
            "engine_load": load,
            "intake_temp": intake_temp,
            "ambient_air_temp": ambient_air_temp,
            "odometer": odometer,
            "temp_cpu": temp_cpu,
            "cpu": cpu,
            "gpu": gpu,
            "thermal_delta": (temp - intake_temp) if temp is not None and intake_temp is not None else None,
        }

        consumed_groups: set[str] = set()
        for rule in rules.get("sensor_rules", []):
            group = str(rule.get("group") or "").strip()
            if rule.get("exclusive_group") and group and group in consumed_groups:
                continue
            if not RecommendationEngine._matches_condition(rule.get("when") or {}, values):
                continue

            risk_cfg = rule.get("risk") or {}
            value = values.get(str(rule.get("value_field") or "").strip(), None)
            if value is None:
                value = values.get(str((risk_cfg.get("type") or "")).strip(), None)
            if value is None and str(risk_cfg.get("type") or "") == "battery":
                value = battery
            if value is None and str(risk_cfg.get("type") or "") == "cooling":
                value = temp
            if value is None and str(risk_cfg.get("type") or "") == "fuel":
                value = fuel

            added = add_risk(
                str(risk_cfg.get("type") or "unknown"),
                str(risk_cfg.get("severity") or "warning"),
                str(risk_cfg.get("message") or "Alerte detectee"),
                value,
                component=rule.get("component"),
            )
            suggestion = rule.get("suggestion")
            if suggestion and added:
                add_suggestion(
                    str(suggestion.get("priority") or "low"),
                    str(suggestion.get("title") or "Action recommandee"),
                    str(suggestion.get("message") or "Verifier le vehicule."),
                )

            if rule.get("exclusive_group") and group:
                consumed_groups.add(group)

        # ── Maintenance par intervalle depuis dernier entretien ─────────────
        def _safe_delta(current: float | None, previous: float | None) -> float | None:
            if current is None or previous is None:
                return None
            try:
                delta = float(current) - float(previous)
            except (TypeError, ValueError):
                return None
            return delta if delta >= 0 else None

        last_oil = maintenance_context.get("last_oil_change_odometer")
        oil_interval = maintenance_context.get("oil_change_interval_km")
        oil_delta = _safe_delta(odometer, last_oil)
        oil_ratio = (float(oil_delta) / float(oil_interval)) if oil_delta is not None and oil_interval and oil_interval > 0 else None

        last_service = maintenance_context.get("last_maintenance_odometer")
        service_interval = maintenance_context.get("maintenance_interval_km")
        service_delta = _safe_delta(odometer, last_service)
        service_ratio = (
            (float(service_delta) / float(service_interval))
            if service_delta is not None and service_interval and service_interval > 0
            else None
        )

        last_parts = maintenance_context.get("last_major_parts_change_odometer")
        parts_interval = maintenance_context.get("major_parts_interval_km")
        parts_delta = _safe_delta(odometer, last_parts)
        parts_ratio = (float(parts_delta) / float(parts_interval)) if parts_delta is not None and parts_interval and parts_interval > 0 else None

        has_maintenance_baseline = bool(maintenance_records) or any(
            maintenance_context.get(k) is not None
            for k in (
                "last_oil_change_odometer",
                "last_maintenance_odometer",
                "last_major_parts_change_odometer",
            )
        )

        values.update(
            {
                "oil_delta": oil_delta,
                "oil_ratio": oil_ratio,
                "service_delta": service_delta,
                "service_ratio": service_ratio,
                "parts_delta": parts_delta,
                "parts_ratio": parts_ratio,
                "has_maintenance_baseline": has_maintenance_baseline,
            }
        )

        for ruleset in ("maintenance_rules", "fallback_rules"):
            consumed_groups = set()
            for rule in rules.get(ruleset, []):
                group = str(rule.get("group") or "").strip()
                if rule.get("exclusive_group") and group and group in consumed_groups:
                    continue
                if not RecommendationEngine._matches_condition(rule.get("when") or {}, values):
                    continue

                risk_cfg = rule.get("risk") or {}
                value_key = str(rule.get("value_field") or "").strip()
                risk_value = values.get(value_key)

                added = add_risk(
                    str(risk_cfg.get("type") or "unknown"),
                    str(risk_cfg.get("severity") or "warning"),
                    str(risk_cfg.get("message") or "Alerte detectee"),
                    risk_value,
                    component=rule.get("component"),
                )
                suggestion = rule.get("suggestion")
                if suggestion and added:
                    add_suggestion(
                        str(suggestion.get("priority") or "low"),
                        str(suggestion.get("title") or "Action recommandee"),
                        str(suggestion.get("message") or "Verifier le vehicule."),
                    )

                if rule.get("exclusive_group") and group:
                    consumed_groups.add(group)

        # ── DTC actifs ───────────────────────────────────────────────────────
        for dtc in active_dtc_events:
            code = str(dtc.get("code") or "DTC_UNKNOWN")
            desc = str(dtc.get("description") or "Code défaut détecté")
            dtc_sev_raw = str(dtc.get("severity") or "warning").lower().strip()
            if dtc_sev_raw not in {"info", "warning", "critical"}:
                dtc_sev_raw = "warning"

            if add_risk("dtc", dtc_sev_raw, f"{code}: {desc}", dtc.get("occurrence_count"), component="dtc", dtc_code=code):
                suggested_action = dtc.get("recommended_action") or f"Diagnostiquer le code {code} et corriger la cause racine."
                add_suggestion(
                    "high" if dtc_sev_raw == "critical" else "medium",
                    f"Code défaut {code}",
                    suggested_action,
                )

        # ── Store raw ML output before any rule adjustment ─────────────────────
        ml_severity   = severity
        ml_risk_score = risk_score

        # ── Rule-based score floor (applied AFTER all sensor checks) ───────────
        # Rules only raise the score/severity; they never lower it.
        # The original ML values are preserved in `rule_override` for transparency.
        _has_critical_rule = any(r["severity"] == "critical" for r in predicted_risks)
        _critical_count    = sum(1 for r in predicted_risks if r["severity"] == "critical")
        _has_warning_rule  = any(r["severity"] == "warning"  for r in predicted_risks)

        score_floor = rules.get("score_floor") or {}
        if _critical_count >= 2:
            risk_score = max(risk_score, float(score_floor.get("critical_count_ge_2", 85.0)))
        elif _has_critical_rule:
            risk_score = max(risk_score, float(score_floor.get("critical_any", 70.0)))
        elif _has_warning_rule:
            risk_score = max(risk_score, float(score_floor.get("warning_any", 40.0)))

        # Promote severity if rule-corrected score crosses thresholds
        severity_promotion = rules.get("severity_promotion") or {}
        critical_min = float(severity_promotion.get("critical_min_score", 70.0))
        warning_min = float(severity_promotion.get("warning_min_score", 40.0))
        if risk_score >= critical_min:
            severity = "critical"
        elif risk_score >= warning_min and severity == "info":
            severity = "warning"

        # Build override metadata after suppression-aware post-processing.
        rule_override: dict | None = None

        if not maintenance_suggestions:
            normal_suggestion = (rules.get("defaults") or {}).get("normal_suggestion") or {}
            add_suggestion(
                str(normal_suggestion.get("priority") or "low"),
                str(normal_suggestion.get("title") or "Etat normal"),
                str(normal_suggestion.get("message") or "Aucune anomalie detectee actuellement."),
            )

        if predicted_risks:
            severity_rank = {"critical": 3, "warning": 2, "info": 1}
            suggestion_priority_rank = {"high": 3, "medium": 2, "low": 1}

            # Keep all active alerts but order them by severity so the UI can
            # display urgent anomalies first without dropping any item.
            predicted_risks.sort(
                key=lambda r: (
                    -severity_rank.get(str(r.get("severity", "info")).lower(), 1),
                    str(r.get("type", "")),
                    str(r.get("message", "")),
                )
            )

            # Same idea for recommendations: high priority first.
            maintenance_suggestions.sort(
                key=lambda s: (
                    -suggestion_priority_rank.get(str(s.get("priority", "low")).lower(), 1),
                    str(s.get("title", "")),
                )
            )

            risk_labels = {
                "battery": "batterie",
                "cooling": "refroidissement",
                "fuel": "carburant",
                "driving_style": "style de conduite",
                "engine_load": "charge moteur",
                "device_cpu_temp": "température CPU",
                "device_cpu_load": "charge CPU",
                "device_gpu_load": "charge GPU",
                "intake": "admission",
                "ambient_heat": "température ambiante",
                "thermal_delta": "écart thermique",
                "driving_speed": "vitesse vehicule",
                "mileage": "kilometrage",
                "dtc": "codes défaut",
            }
            alert_domains = []
            for risk in predicted_risks:
                domain = risk_labels.get(risk.get("type", ""), str(risk.get("type", "alerte")))
                if domain and domain not in alert_domains:
                    alert_domains.append(domain)

            if risk_score >= 75 or severity == "critical":
                priority = "high"
            elif risk_score >= 40 or severity == "warning":
                priority = "medium"
            else:
                priority = "low"

            if len(alert_domains) == 1:
                summary = f"Alerte détectée: {alert_domains[0]}."
            else:
                summary = f"Alertes détectées: {', '.join(alert_domains)}."

            # Show all active anomalies (after suppression of already-resolved ones)
            # and keep priority order for quick operational reading.
            key_messages: list[str] = []
            for risk in predicted_risks:
                msg = str(risk.get("message") or "").strip()
                if msg and msg not in key_messages:
                    key_messages.append(msg)

            next_action = f"Anomalies prioritaires: {' | '.join(key_messages)}" if key_messages else ""
        else:
            # Coherence guardrail: with no visible anomalies, never expose
            # warning/critical severity. Keep score in a low-safe range.
            severity = "info"
            risk_score = min(risk_score, 30.0)
            priority = "low"
            summary = "Aucune anomalie détectée."
            next_action = ""

        _score_changed = round(risk_score, 2) != round(ml_risk_score, 2)
        _severity_changed = severity != ml_severity
        if _score_changed or _severity_changed:
            triggered = []
            if _critical_count >= 2:
                triggered.append("double_critical_floor_85")
            elif _has_critical_rule:
                triggered.append("critical_floor_70")
            elif _has_warning_rule:
                triggered.append("warning_floor_40")
            if suppressed_alerts and not predicted_risks:
                triggered.append("all_alerts_suppressed_force_info")
            elif not predicted_risks:
                triggered.append("no_visible_risk_force_info")
            if _severity_changed:
                triggered.append(f"severity_promoted_{ml_severity}_to_{severity}")
            rule_override = {
                "applied": True,
                "triggered_rules": triggered,
                "ml_severity": ml_severity,
                "ml_risk_score": round(ml_risk_score, 2),
                "adjusted_severity": severity,
                "adjusted_risk_score": round(risk_score, 2),
            }
        else:
            rule_override = {
                "applied": False,
                "ml_severity": ml_severity,
                "ml_risk_score": round(ml_risk_score, 2),
            }

        maintenance_required = any(r["severity"] in {"warning", "critical"} for r in predicted_risks)
        preventive_only = (not maintenance_required) and any(r["severity"] == "info" for r in predicted_risks)
        maintenance_reasons = [str(r["message"]) for r in predicted_risks[:5]]

        if any(r["severity"] == "critical" for r in predicted_risks):
            maintenance_status = {
                "maintenance_required": True,
                "maintenance_type": "urgent",
                "priority": "high",
                "summary": "Maintenance urgente requise.",
                "reasons": maintenance_reasons,
            }
        elif maintenance_required:
            maintenance_status = {
                "maintenance_required": True,
                "maintenance_type": "planned",
                "priority": "medium",
                "summary": "Maintenance recommandee prochainement.",
                "reasons": maintenance_reasons,
            }
        elif preventive_only:
            maintenance_status = {
                "maintenance_required": False,
                "maintenance_type": "preventive",
                "priority": "low",
                "summary": "Entretien preventif conseille.",
                "reasons": maintenance_reasons,
            }
        else:
            maintenance_status = {
                "maintenance_required": False,
                "maintenance_type": "none",
                "priority": "low",
                "summary": "Aucune maintenance necessaire actuellement.",
                "reasons": [],
            }

        prediction["predicted_risks"] = predicted_risks
        prediction["maintenance_filter"] = {
            "applied": bool(suppressed_alerts),
            "suppressed_alerts": suppressed_alerts,
        }
        prediction["maintenance_status"] = maintenance_status
        prediction["maintenance_suggestions"] = maintenance_suggestions
        prediction["predicted_risk_score"] = round(risk_score, 2)
        prediction["predicted_severity"]   = severity
        prediction["rule_override"]        = rule_override
        prediction["ai_insights"] = {
            "summary": summary,
            "priority": priority,
            "next_action": next_action,
        }
        return prediction
